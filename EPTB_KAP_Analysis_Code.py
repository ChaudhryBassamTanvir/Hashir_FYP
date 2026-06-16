"""
EPTB KAP Study — Complete Data Analysis Code
Project: Assessment of Knowledge, Attitude and Practice regarding
         Extra-Pulmonary Tuberculosis (EPTB) among Community Members

Requirements:
    pip install pandas openpyxl scipy

Usage:
    1. Place your Excel file in the same folder as this script
    2. Change FILE_NAME below if your file is named differently
    3. Run: python EPTB_KAP_Analysis_Code.py
    4. Output: EPTB_KAP_Analysis.xlsx (all 6 tables)
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from scipy.stats import chi2_contingency

# ── CONFIG ────────────────────────────────────────────────────────────────────
FILE_NAME   = 'Untitled_form__Responses_.xlsx'
OUTPUT_NAME = 'EPTB_KAP_Analysis.xlsx'

# ── LOAD DATA ─────────────────────────────────────────────────────────────────
df = pd.read_excel(FILE_NAME)
N  = len(df)
print(f"Data loaded: {N} rows, {len(df.columns)} columns")

# ── HELPER FUNCTIONS ──────────────────────────────────────────────────────────
def hdr(ws, row, col, val):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(bold=True, color='FFFFFF', size=11, name='Arial')
    c.fill      = PatternFill('solid', start_color='1F4E79')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def cell(ws, row, col, val, bold=False, align='center'):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(bold=bold, name='Arial', size=10)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)

def zebra(ws, row, ncols):
    """Light blue zebra stripe on even rows"""
    if row % 2 == 0:
        for c in range(1, ncols+1):
            ws.cell(row, c).fill = PatternFill('solid', start_color='EBF5FB')

def section_row(ws, row, ncols):
    """Dark header row for sub-sections inside Table 1"""
    for c in range(1, ncols+1):
        ws.cell(row, c).fill      = PatternFill('solid', start_color='D6E4F0')
        ws.cell(row, c).font      = Font(bold=True, name='Arial', size=10)

def border_range(ws, r1, c1, r2, c2):
    thin = Side(style='thin')
    for r in range(r1, r2+1):
        for c in range(c1, c2+1):
            ws.cell(r, c).border = Border(
                left=thin, right=thin, top=thin, bottom=thin)

def pct(n, total):
    return f"{n} ({n/total*100:.1f}%)"

# ── KAP SCORING ───────────────────────────────────────────────────────────────
# Knowledge (13 items) — correct answers defined below
know_correct = {
    df.columns[6]:  'Yes',                  # Heard about EPTB
    df.columns[7]:  'Other parts of body',  # Where EPTB occurs (correct = other parts)
    df.columns[8]:  'Yes',                  # Bacterium causes TB
    df.columns[9]:  'Yes',                  # Know body parts
    df.columns[10]: 'Yes',                  # Know symptoms
    df.columns[11]: 'Yes',                  # TB communicable
    df.columns[12]: 'Yes',                  # Living with patient = transmission
    df.columns[13]: 'Yes',                  # EPTB preventable
    df.columns[14]: 'Yes',                  # EPTB curable
    df.columns[15]: 'Yes',                  # Drug therapy cures EPTB
    df.columns[16]: 'No',                   # Cost (correct = No, treatment is free)
    df.columns[17]: 'Yes',                  # Complete full course
    df.columns[18]: 'Yes',                  # Vaccines help
}
df['know_score'] = sum((df[c] == v).astype(int) for c, v in know_correct.items())
df['knowledge']  = (df['know_score'] >= 0.6 * 13).map(
    {True: 'Adequate', False: 'Inadequate'})

# Attitude (7 items) — positive responses defined below
att_correct = {
    df.columns[19]: 'Yes',  # Feel hope
    df.columns[20]: 'Yes',  # Tell others
    df.columns[21]: 'No',   # Stay away (No = positive attitude)
    df.columns[22]: 'Yes',  # EPTB serious problem
    df.columns[23]: 'Yes',  # Health education prevents spread
    df.columns[24]: 'Yes',  # Allow child to marry cured patient
    df.columns[27]: 'No',   # EPTB less dangerous (No = correct)
}
df['att_score'] = sum((df[c] == v).astype(int) for c, v in att_correct.items())
df['attitude']  = (df['att_score'] >= 0.6 * 7).map(
    {True: 'Positive', False: 'Negative'})

# Practice (8 items) — all Yes = good practice
prac_cols = [df.columns[i] for i in [25, 26, 28, 29, 30, 31, 32, 33]]
df['prac_score'] = sum((df[c] == 'Yes').astype(int) for c in prac_cols)
df['practice']   = (df['prac_score'] >= 0.6 * 8).map(
    {True: 'Good', False: 'Poor'})

print(f"Adequate knowledge : {(df['knowledge']=='Adequate').sum()} ({(df['knowledge']=='Adequate').sum()/N*100:.1f}%)")
print(f"Positive attitude  : {(df['attitude']=='Positive').sum()} ({(df['attitude']=='Positive').sum()/N*100:.1f}%)")
print(f"Good practice      : {(df['practice']=='Good').sum()} ({(df['practice']=='Good').sum()/N*100:.1f}%)")

# ── CREATE WORKBOOK ───────────────────────────────────────────────────────────
wb = Workbook()

# ── TABLE 1: DEMOGRAPHICS ─────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = 'Table 1 - Demographics'
ws1.column_dimensions['A'].width = 38
ws1.column_dimensions['B'].width = 22
ws1.column_dimensions['C'].width = 22
ws1.row_dimensions[1].height = 30

hdr(ws1, 1, 1, 'Characteristics')
hdr(ws1, 1, 2, 'Frequency')
hdr(ws1, 1, 3, 'Percentage (%)')

demo_rows = [
    ('Total participants', N, '100%'),
    ('Gender', '', ''),
]
for v, n in df['Gender'].value_counts().items():
    demo_rows.append((f'  {v}', n, f'{n/N*100:.1f}%'))
demo_rows.append(('Age group', '', ''))
for v in ['Below 15', '15 to 20', '21 to 25', '26 to 30', 'Above 30']:
    n = (df['Age'] == v).sum()
    demo_rows.append((f'  {v}', n, f'{n/N*100:.1f}%'))
demo_rows.append(('Education', '', ''))
for v, n in df['Education '].value_counts().items():
    demo_rows.append((f'  {v}', n, f'{n/N*100:.1f}%'))
demo_rows.append(('Medical background', '', ''))
for v, n in df['Any Medical background '].value_counts().items():
    demo_rows.append((f'  {v}', n, f'{n/N*100:.1f}%'))
demo_rows.append(('Health education on EPTB received', '', ''))
for v, n in df['Health education on Extra-pulmonary tuberculosis '].value_counts().items():
    demo_rows.append((f'  {v}', n, f'{n/N*100:.1f}%'))

for i, (a, b, c) in enumerate(demo_rows, start=2):
    is_section = (b == '' and a != 'Total participants')
    cell(ws1, i, 1, a, bold=is_section, align='left')
    cell(ws1, i, 2, b, bold=is_section)
    cell(ws1, i, 3, c, bold=is_section)
    if is_section:
        section_row(ws1, i, 3)
    else:
        zebra(ws1, i, 3)

border_range(ws1, 1, 1, len(demo_rows)+1, 3)

# ── TABLE 2: KNOWLEDGE ────────────────────────────────────────────────────────
ws2 = wb.create_sheet('Table 2 - Knowledge')
ws2.column_dimensions['A'].width = 52
ws2.column_dimensions['B'].width = 16
ws2.column_dimensions['C'].width = 16
ws2.column_dimensions['D'].width = 16
ws2.column_dimensions['E'].width = 16
ws2.row_dimensions[1].height = 35

for col, label in enumerate(
    ['Knowledge Question', 'Yes', 'No', "Don't Know / Maybe", 'Total'], start=1):
    hdr(ws2, 1, col, label)

know_items = [
    ('Have you ever heard about EPTB?',                             df.columns[6]),
    ('EPTB occurs in (correct = Other parts of body)',              df.columns[7]),
    ('A bacterium is the cause of TB?',                            df.columns[8]),
    ('Do you know body parts where EPTB occurs?',                  df.columns[9]),
    ('Do you know the symptoms of EPTB?',                          df.columns[10]),
    ('TB is communicable from person to person?',                  df.columns[11]),
    ('Living with TB patient is a mode of transmission?',          df.columns[12]),
    ('EPTB is a preventable disease?',                             df.columns[13]),
    ('EPTB is curable?',                                           df.columns[14]),
    ('EPTB specific drug therapy is a method to cure EPTB?',       df.columns[15]),
    ('Do you know the cost of EPTB treatment?',                    df.columns[16]),
    ('Necessary to complete the full course of EPTB treatment?',   df.columns[17]),
    ('Can vaccines help prevent EPTB?',                            df.columns[18]),
]

for i, (label, col) in enumerate(know_items, start=2):
    vals = df[col].value_counts()
    yes  = vals.get('Yes', 0)
    no   = vals.get('No', 0)
    dk   = vals.get("Don't no", vals.get('Maybe', 0))
    cell(ws2, i, 1, label, align='left')
    cell(ws2, i, 2, pct(yes, N))
    cell(ws2, i, 3, pct(no, N))
    cell(ws2, i, 4, pct(dk, N) if dk else '-')
    cell(ws2, i, 5, N)
    zebra(ws2, i, 5)

border_range(ws2, 1, 1, len(know_items)+1, 5)

# ── TABLE 3: ATTITUDE ─────────────────────────────────────────────────────────
ws3 = wb.create_sheet('Table 3 - Attitude')
ws3.column_dimensions['A'].width = 52
ws3.column_dimensions['B'].width = 16
ws3.column_dimensions['C'].width = 16
ws3.column_dimensions['D'].width = 16
ws3.row_dimensions[1].height = 30

for col, label in enumerate(['Attitude Question', 'Yes', 'No', 'Total'], start=1):
    hdr(ws3, 1, col, label)

att_items = [
    ('Would you feel hope if found to have EPTB?',          df.columns[19]),
    ('Would you tell others if you develop EPTB?',          df.columns[20]),
    ('Do you stay away from people with EPTB?',             df.columns[21]),
    ('EPTB is a serious health problem in community?',      df.columns[22]),
    ('Health education can prevent the spread of EPTB?',   df.columns[23]),
    ('Allow your child to marry a cured EPTB patient?',    df.columns[24]),
    ('EPTB is less dangerous than pulmonary TB?',           df.columns[27]),
]

for i, (label, col) in enumerate(att_items, start=2):
    vals = df[col].value_counts()
    cell(ws3, i, 1, label, align='left')
    cell(ws3, i, 2, pct(vals.get('Yes', 0), N))
    cell(ws3, i, 3, pct(vals.get('No', 0), N))
    cell(ws3, i, 4, N)
    zebra(ws3, i, 4)

border_range(ws3, 1, 1, len(att_items)+1, 4)

# ── TABLE 4: PRACTICE ─────────────────────────────────────────────────────────
ws4 = wb.create_sheet('Table 4 - Practice')
ws4.column_dimensions['A'].width = 52
ws4.column_dimensions['B'].width = 16
ws4.column_dimensions['C'].width = 16
ws4.column_dimensions['D'].width = 16
ws4.row_dimensions[1].height = 30

for col, label in enumerate(['Practice Question', 'Yes', 'No', 'Total'], start=1):
    hdr(ws4, 1, col, label)

prac_items = [
    ('Have you ever got health education about EPTB?',                    df.columns[25]),
    ('Have you ever been screened for EPTB?',                             df.columns[26]),
    ('If you have EPTB, do you consult a health worker?',                 df.columns[28]),
    ('If you develop EPTB symptoms, will you go to health facility?',     df.columns[29]),
    ('Have you advised EPTB patients to take their drugs properly?',      df.columns[30]),
    ('Have you covered your mouth during coughing?',                      df.columns[31]),
    ('Do you follow hygiene measures to avoid EPTB?',                     df.columns[32]),
    ('Early diagnosis and treatment can prevent further complications?',  df.columns[33]),
]

for i, (label, col) in enumerate(prac_items, start=2):
    vals = df[col].value_counts()
    cell(ws4, i, 1, label, align='left')
    cell(ws4, i, 2, pct(vals.get('Yes', 0), N))
    cell(ws4, i, 3, pct(vals.get('No', 0), N))
    cell(ws4, i, 4, N)
    zebra(ws4, i, 4)

border_range(ws4, 1, 1, len(prac_items)+1, 4)

# ── TABLE 5: CHI-SQUARE ASSOCIATIONS ─────────────────────────────────────────
ws5 = wb.create_sheet('Table 5 - Chi-square')
ws5.column_dimensions['A'].width = 22
ws5.column_dimensions['B'].width = 30
ws5.column_dimensions['C'].width = 14
ws5.column_dimensions['D'].width = 14
ws5.row_dimensions[1].height = 30

for col, label in enumerate(
    ['Variable', 'KAP Outcome', 'Chi-square', 'p-value'], start=1):
    hdr(ws5, 1, col, label)

combos = [
    ('Gender',           'Gender',                        'knowledge', 'Adequate / Inadequate Knowledge'),
    ('Gender',           'Gender',                        'attitude',  'Positive / Negative Attitude'),
    ('Gender',           'Gender',                        'practice',  'Good / Poor Practice'),
    ('Age group',        'Age',                           'knowledge', 'Adequate / Inadequate Knowledge'),
    ('Age group',        'Age',                           'attitude',  'Positive / Negative Attitude'),
    ('Age group',        'Age',                           'practice',  'Good / Poor Practice'),
    ('Education',        'Education ',                    'knowledge', 'Adequate / Inadequate Knowledge'),
    ('Education',        'Education ',                    'attitude',  'Positive / Negative Attitude'),
    ('Education',        'Education ',                    'practice',  'Good / Poor Practice'),
    ('Medical background','Any Medical background ',      'knowledge', 'Adequate / Inadequate Knowledge'),
    ('Medical background','Any Medical background ',      'attitude',  'Positive / Negative Attitude'),
    ('Medical background','Any Medical background ',      'practice',  'Good / Poor Practice'),
]

for i, (label, col, kap, kap_label) in enumerate(combos, start=2):
    ct = pd.crosstab(df[col], df[kap])
    if ct.shape[0] >= 2 and ct.shape[1] >= 2:
        chi2, p, _, _ = chi2_contingency(ct)
        chi2_str = f"{chi2:.3f}"
        p_str    = f"{p:.3f}{' *' if p < 0.05 else ''}"
    else:
        chi2_str, p_str = 'N/A', 'N/A'
    cell(ws5, i, 1, label, bold=True, align='left')
    cell(ws5, i, 2, kap_label, align='left')
    cell(ws5, i, 3, chi2_str)
    cell(ws5, i, 4, p_str)
    zebra(ws5, i, 4)

border_range(ws5, 1, 1, len(combos)+1, 4)
note = ws5.cell(row=len(combos)+3, column=1, value='* p < 0.05 = statistically significant')
note.font = Font(italic=True, size=9, name='Arial')

# ── TABLE 6: KAP SUMMARY ──────────────────────────────────────────────────────
ws6 = wb.create_sheet('Table 6 - KAP Summary')
ws6.column_dimensions['A'].width = 20
ws6.column_dimensions['B'].width = 28
ws6.column_dimensions['C'].width = 20
ws6.column_dimensions['D'].width = 14
ws6.row_dimensions[1].height = 30

for col, label in enumerate(['KAP Domain', 'Category', 'n (%)', 'Total'], start=1):
    hdr(ws6, 1, col, label)

kap_summary = [
    ('Knowledge', 'Adequate (≥60% correct)',    (df['knowledge'] == 'Adequate').sum()),
    ('Knowledge', 'Inadequate (<60% correct)',  (df['knowledge'] == 'Inadequate').sum()),
    ('Attitude',  'Positive (≥60% correct)',    (df['attitude']  == 'Positive').sum()),
    ('Attitude',  'Negative (<60% correct)',    (df['attitude']  == 'Negative').sum()),
    ('Practice',  'Good (≥60% correct)',        (df['practice']  == 'Good').sum()),
    ('Practice',  'Poor (<60% correct)',        (df['practice']  == 'Poor').sum()),
]

for i, (domain, cat, n) in enumerate(kap_summary, start=2):
    cell(ws6, i, 1, domain, bold=True)
    cell(ws6, i, 2, cat, align='left')
    cell(ws6, i, 3, pct(n, N))
    cell(ws6, i, 4, N)
    zebra(ws6, i, 4)

border_range(ws6, 1, 1, len(kap_summary)+1, 4)

# ── SAVE ──────────────────────────────────────────────────────────────────────
wb.save(OUTPUT_NAME)
print(f"\nDone! File saved as: {OUTPUT_NAME}")
print("6 sheets: Demographics | Knowledge | Attitude | Practice | Chi-square | KAP Summary")